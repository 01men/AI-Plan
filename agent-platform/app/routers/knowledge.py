"""知识库：NAS 空间与文档（R4-3 文件上传 + 自动解析拆分）"""
import csv
import io
import json
import re
import shutil
import sqlite3
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse

from app.access import can_access_document, require_document
from app.database import DB_PATH
from app.routers.auth import audit, db_conn, get_current_person

router = APIRouter(prefix="/api/knowledge", tags=["knowledge"])

UPLOAD_ROOT = DB_PATH.parent / "uploads"          # data/uploads/space_{sid}/
KNOWLEDGE_DB_DIR = DB_PATH.parent / "knowledge"   # data/knowledge/doc_{id}.db

ALLOWED_EXT = (
    ".txt", ".md", ".docx", ".pdf", ".csv", ".json", ".xlsx", ".xls", ".html", ".htm",
)
UPLOAD_TIERS = ("boss", "coach", "backbone", "developer")
MAX_EXCEL_SHEETS = 50
MAX_EXCEL_ROWS_PER_SHEET = 100_000
MAX_EXCEL_COLUMNS = 200
MAX_EXCEL_CELLS = 1_000_000


@router.get("/spaces")
def list_spaces(conn=Depends(db_conn), person=Depends(get_current_person)):
    out = []
    for r in conn.execute("SELECT * FROM knowledge_spaces ORDER BY id"):
        d = dict(r)
        docs = conn.execute("SELECT level FROM documents WHERE space_id=?", (r["id"],)).fetchall()
        d["doc_count"] = sum(
            1 for doc in docs if can_access_document(person, doc["level"], r["dept_name"])
        )
        out.append(d)
    return out


@router.get("/documents")
def list_documents(space_id: int = None, level: str = None, conn=Depends(db_conn),
                   person=Depends(get_current_person)):
    sql = ("SELECT d.*,s.name space_name,s.dept_name space_dept FROM documents d "
           "JOIN knowledge_spaces s ON s.id=d.space_id")
    cond, args = [], []
    if space_id:
        cond.append("d.space_id=?")
        args.append(space_id)
    if level:
        cond.append("d.level=?")
        args.append(level)
    if cond:
        sql += " WHERE " + " AND ".join(cond)
    sql += " ORDER BY d.id"
    return [dict(r) for r in conn.execute(sql, args)
            if can_access_document(person, r["level"], r["space_dept"])]


@router.get("/business-data")
def list_business_data(business_type: str = None, q: str = None, limit: int = 20,
                       offset: int = 0, conn=Depends(db_conn),
                       person=Depends(get_current_person)):
    """默认 1000 条制造业务展示数据，支持分类/关键词分页查询。"""
    del person  # 所有已登录员工均可读取系统演示数据
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    cond, args = [], []
    if business_type:
        cond.append("business_type=?")
        args.append(business_type)
    if q and q.strip():
        keyword = f"%{q.strip()}%"
        cond.append(
            "(record_no LIKE ? OR customer LIKE ? OR product_code LIKE ? OR "
            "product_name LIKE ? OR status LIKE ? OR department LIKE ?)"
        )
        args.extend([keyword] * 6)
    where = (" WHERE " + " AND ".join(cond)) if cond else ""
    total = conn.execute("SELECT COUNT(*) c FROM business_records" + where, args).fetchone()["c"]
    rows = [dict(row) for row in conn.execute(
        "SELECT * FROM business_records" + where
        + " ORDER BY business_date DESC,record_no LIMIT ? OFFSET ?",
        args + [limit, offset],
    )]
    summary = [dict(row) for row in conn.execute(
        "SELECT business_type,COUNT(*) count,ROUND(SUM(amount),2) amount,"
        "ROUND(AVG(metric_value),1) avg_metric FROM business_records "
        "GROUP BY business_type ORDER BY business_type"
    )]
    return {"total": total, "limit": limit, "offset": offset,
            "summary": summary, "items": rows}


@router.post("/documents")
def create_document(body: dict = Body(...), conn=Depends(db_conn),
                    person=Depends(get_current_person)):
    if person["tier"] not in UPLOAD_TIERS:
        raise HTTPException(403, "仅业务骨干、教练团、高管或开发者可登记文档")
    title = (body.get("title") or "").strip()
    space_id = body.get("space_id")
    if not title or not space_id:
        raise HTTPException(400, "title 与 space_id 必填")
    space = conn.execute("SELECT * FROM knowledge_spaces WHERE id=?", (space_id,)).fetchone()
    if not space:
        raise HTTPException(404, "知识空间不存在")
    level = str(body.get("level", "L1")).upper()
    if level not in ("L1", "L2", "L3", "L4"):
        raise HTTPException(422, "密级仅支持 L1/L2/L3/L4")
    if level == "L4" and person["tier"] not in ("boss", "coach", "backbone"):
        raise HTTPException(403, "仅高管、教练团或业务骨干可登记 L4 文档")
    did = conn.execute(
        "INSERT INTO documents(space_id,title,level,tags,uploaded_by,created_at) VALUES(?,?,?,?,?,?)",
        (space_id, title, level, body.get("tags", ""), person["name"],
         datetime.now().isoformat(timespec="seconds"))).lastrowid
    conn.commit()
    audit(conn, person["name"], "上传文档", title, f"空间 #{space_id}")
    return dict(conn.execute("SELECT * FROM documents WHERE id=?", (did,)).fetchone())


# ---------------- R4-3 上传与自动解析 ----------------

def _decode(raw: bytes) -> str:
    for enc in ("utf-8", "gbk"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _docx_to_text(raw: bytes) -> str:
    """docx 本质是 zip：解出 word/document.xml 正则提取文本"""
    import io
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
        xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    except Exception:
        raise HTTPException(422, "DOCX 解析失败：文件损坏或不是有效的 Word 文档")
    xml = re.sub(r"</w:p>", "\n", xml)       # 段落换行
    xml = re.sub(r"<w:tab[^>]*/>", "\t", xml)
    text = re.sub(r"<[^>]+>", "", xml)        # 去全部标签
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _pdftotext_bin():
    """定位系统 pdftotext：PATH 优先，兜底 Git for Windows 自带路径"""
    w = shutil.which("pdftotext")
    if w:
        return w
    for c in (r"C:\Program Files\Git\mingw64\bin\pdftotext.exe",
              "/mingw64/bin/pdftotext"):
        if Path(c).exists():
            return c
    return None


def _pdf_to_text(raw: bytes, work_dir: Path, doc_id: int) -> str:
    """pdf 走系统 pdftotext；缺失或执行失败给中文错误"""
    bin_path = _pdftotext_bin()
    if not bin_path:
        raise HTTPException(422, "PDF 解析失败：系统未找到 pdftotext（请安装 poppler 或使用 Git 自带版本）")
    src = work_dir / f"doc_{doc_id}_src.pdf"
    dst = work_dir / f"doc_{doc_id}_pdf.txt"
    src.write_bytes(raw)
    try:
        r = subprocess.run([bin_path, "-enc", "UTF-8", str(src), str(dst)],
                           capture_output=True, timeout=60)
    except Exception:
        raise HTTPException(422, "PDF 解析失败：pdftotext 执行异常")
    if r.returncode != 0 or not dst.exists():
        raise HTTPException(422, "PDF 解析失败：pdftotext 无法处理该文件（可能为扫描件或已加密）")
    text = dst.read_text(encoding="utf-8", errors="ignore").strip()
    src.unlink(missing_ok=True)
    dst.unlink(missing_ok=True)
    if not text:
        raise HTTPException(422, "PDF 解析失败：未提取到文本内容（可能为纯图片扫描件）")
    return text


def _sanitize_col(name, idx):
    name = re.sub(r"\W+", "_", str(name or "").strip()).strip("_")
    return name or f"col_{idx}"


def _normalize_cell(value):
    """将 Excel/JSON 单元格转为 SQLite 与 CSV 都可稳定保存的标量。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, bool):
        return 1 if value else 0
    return value


def _sqlite_type(values):
    nonempty = [v for v in values if v not in (None, "")]
    if not nonempty:
        return "TEXT"
    if all(isinstance(v, (bool, int)) and not isinstance(v, float) for v in nonempty):
        return "INTEGER"
    if all(isinstance(v, (bool, int, float)) for v in nonempty):
        return "REAL"
    return "TEXT"


def _unique_names(names, prefix="col"):
    out, seen = [], {}
    for index, name in enumerate(names, 1):
        base = _sanitize_col(name, index) or f"{prefix}_{index}"
        seen[base] = seen.get(base, 0) + 1
        out.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return out


def _rows_to_sqlite(doc_id: int, headers, rows):
    """表格类数据写入独立 SQLite（每文档一个表 data_rows），返回 db 路径"""
    KNOWLEDGE_DB_DIR.mkdir(parents=True, exist_ok=True)
    db_file = KNOWLEDGE_DB_DIR / f"doc_{doc_id}.db"
    db_file.unlink(missing_ok=True)
    cols = _unique_names(headers)
    types = [_sqlite_type([row[index] for row in rows if index < len(row)])
             for index in range(len(cols))]
    kconn = sqlite3.connect(str(db_file))
    try:
        kconn.execute(
            f"CREATE TABLE data_rows("
            f"{', '.join(f'\"{col}\" {typ}' for col, typ in zip(cols, types))})"
        )
        kconn.executemany(
            f"INSERT INTO data_rows VALUES({', '.join('?' * len(cols))})",
            [[_normalize_cell(v) for v in row] for row in rows])
        kconn.commit()
    finally:
        kconn.close()
    return db_file, cols


def _excel_parse(raw: bytes, ext: str):
    """解析 xlsx/xls 的全部非空工作表，保留数值/日期类型。"""
    datasets = []
    total_cells = 0
    try:
        if ext == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            if len(workbook.worksheets) > MAX_EXCEL_SHEETS:
                workbook.close()
                raise HTTPException(422, f"Excel 工作表不能超过 {MAX_EXCEL_SHEETS} 个")
            sheets = []
            for sheet in workbook.worksheets:
                if sheet.max_row > MAX_EXCEL_ROWS_PER_SHEET or sheet.max_column > MAX_EXCEL_COLUMNS:
                    workbook.close()
                    raise HTTPException(
                        422, f"工作表「{sheet.title}」超过 {MAX_EXCEL_ROWS_PER_SHEET} 行"
                        f"或 {MAX_EXCEL_COLUMNS} 列限制"
                    )
                total_cells += sheet.max_row * sheet.max_column
                if total_cells > MAX_EXCEL_CELLS:
                    workbook.close()
                    raise HTTPException(422, f"Excel 总单元格不能超过 {MAX_EXCEL_CELLS} 个")
                sheets.append((
                    sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]
                ))
            workbook.close()
        else:
            import xlrd

            book = xlrd.open_workbook(file_contents=raw, on_demand=True)
            if book.nsheets > MAX_EXCEL_SHEETS:
                book.release_resources()
                raise HTTPException(422, f"Excel 工作表不能超过 {MAX_EXCEL_SHEETS} 个")
            sheets = []
            for sheet in book.sheets():
                if sheet.nrows > MAX_EXCEL_ROWS_PER_SHEET or sheet.ncols > MAX_EXCEL_COLUMNS:
                    book.release_resources()
                    raise HTTPException(
                        422, f"工作表「{sheet.name}」超过 {MAX_EXCEL_ROWS_PER_SHEET} 行"
                        f"或 {MAX_EXCEL_COLUMNS} 列限制"
                    )
                total_cells += sheet.nrows * sheet.ncols
                if total_cells > MAX_EXCEL_CELLS:
                    book.release_resources()
                    raise HTTPException(422, f"Excel 总单元格不能超过 {MAX_EXCEL_CELLS} 个")
                values = []
                for row_index in range(sheet.nrows):
                    row = []
                    for col_index in range(sheet.ncols):
                        cell = sheet.cell(row_index, col_index)
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            value = xlrd.xldate_as_datetime(value, book.datemode)
                        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            value = bool(value)
                        elif cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                            value = int(value)
                        row.append(value)
                    values.append(row)
                sheets.append((sheet.name, values))
            book.release_resources()
    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(500, "Excel 解析组件未安装，请重新运行启动脚本安装 requirements.txt")
    except Exception:
        raise HTTPException(422, "Excel 解析失败：文件损坏、加密或格式与扩展名不一致")

    for sheet_name, raw_rows in sheets:
        rows = [row for row in raw_rows if any(v not in (None, "") for v in row)]
        if not rows:
            continue
        width = max(len(row) for row in rows)
        rows = [row + [""] * (width - len(row)) for row in rows]
        headers = [
            str(value).strip() if value not in (None, "") else f"列{index + 1}"
            for index, value in enumerate(rows[0])
        ]
        datasets.append({"sheet_name": sheet_name, "headers": headers, "rows": rows[1:]})
    if not datasets:
        raise HTTPException(422, "Excel 解析失败：工作簿没有可读取的数据")
    return datasets


def _excel_to_sqlite_csv(doc_id: int, datasets):
    """每个工作表写一个 SQLite 表，并逐表生成 UTF-8 CSV。"""
    KNOWLEDGE_DB_DIR.mkdir(parents=True, exist_ok=True)
    db_file = KNOWLEDGE_DB_DIR / f"doc_{doc_id}.db"
    csv_dir = KNOWLEDGE_DB_DIR / f"doc_{doc_id}_csv"
    db_file.unlink(missing_ok=True)
    if csv_dir.exists():
        shutil.rmtree(csv_dir)
    csv_dir.mkdir(parents=True)
    conn = sqlite3.connect(str(db_file))
    metadata, used_tables = [], set()
    try:
        conn.execute(
            "CREATE TABLE sheet_index("
            "sheet_name TEXT,table_name TEXT UNIQUE,row_count INTEGER,column_count INTEGER,"
            "csv_path TEXT)"
        )
        for sheet_index, dataset in enumerate(datasets, 1):
            base = "sheet_" + _sanitize_col(dataset["sheet_name"], sheet_index)
            table_name = base
            suffix = 2
            while table_name in used_tables:
                table_name, suffix = f"{base}_{suffix}", suffix + 1
            used_tables.add(table_name)
            headers = _unique_names(dataset["headers"])
            rows = dataset["rows"]
            types = [_sqlite_type([row[index] for row in rows if index < len(row)])
                     for index in range(len(headers))]
            conn.execute(
                f'CREATE TABLE "{table_name}"('
                + ", ".join(f'"{col}" {typ}' for col, typ in zip(headers, types)) + ")"
            )
            if rows:
                conn.executemany(
                    f'INSERT INTO "{table_name}" VALUES({",".join("?" * len(headers))})',
                    [[_normalize_cell(value) for value in row[:len(headers)]] for row in rows],
                )
            csv_path = csv_dir / f"{table_name}.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(headers)
                writer.writerows([[_normalize_cell(value) for value in row[:len(headers)]]
                                  for row in rows])
            item = {
                "sheet_name": dataset["sheet_name"], "table_name": table_name,
                "row_count": len(rows), "column_count": len(headers),
                "columns": headers, "csv_path": str(csv_path),
            }
            metadata.append(item)
            conn.execute(
                "INSERT INTO sheet_index(sheet_name,table_name,row_count,column_count,csv_path)"
                " VALUES(?,?,?,?,?)",
                (item["sheet_name"], table_name, len(rows), len(headers), str(csv_path)),
            )
        conn.commit()
    finally:
        conn.close()
    return db_file, metadata


def _csv_parse(text: str):
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = [r for r in csv.reader(text.splitlines(), dialect) if any(c.strip() for c in r)]
    if not rows:
        raise HTTPException(422, "CSV 解析失败：文件内容为空")
    headers, data = rows[0], rows[1:]
    width = len(headers)
    data = [r + [""] * (width - len(r)) if len(r) < width else r[:width] for r in data]
    return headers, data


def _json_parse(text: str):
    try:
        obj = json.loads(text)
    except Exception:
        raise HTTPException(422, "JSON 解析失败：不是合法的 JSON 文件")
    if isinstance(obj, list) and obj and all(isinstance(x, dict) for x in obj):
        records = obj
    elif isinstance(obj, dict):
        records = None
        for v in obj.values():  # 找第一个 dict 列表字段
            if isinstance(v, list) and v and all(isinstance(x, dict) for x in v):
                records = v
                break
        if records is None:     # 兜底：单条键值对一行
            records = [obj]
    else:
        raise HTTPException(422, "JSON 解析失败：需为对象数组或包含对象数组的 JSON")
    headers = []
    for r in records:
        for k in r.keys():
            if k not in headers:
                headers.append(k)
    rows = [[json.dumps(r.get(h), ensure_ascii=False) if isinstance(r.get(h), (dict, list))
             else r.get(h, "") for h in headers] for r in records]
    return headers, rows


def _md_table(headers, rows):
    """列名 + 前 5 行渲染为 Markdown 表（csv/json 的 .md 摘要）"""
    lines = ["| " + " | ".join(str(h) for h in headers) + " |",
             "|" + " --- |" * len(headers)]
    for r in rows[:5]:
        lines.append("| " + " | ".join(str(v) for v in r) + " |")
    return "\n".join(lines)


def _clean_html(text: str) -> str:
    """网页类清洗：去 script/style 及其内容"""
    text = re.sub(r"<script[\s\S]*?</script>", "", text, flags=re.I)
    text = re.sub(r"<style[\s\S]*?</style>", "", text, flags=re.I)
    return text


def _html_to_text(html: str) -> str:
    text = re.sub(r"<(br|/p|/div|/li|/h[1-6]|/tr)[^>]*>", "\n", html, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s*\n+", "\n\n", text)
    return text.strip()


def _chunk_text(text: str, size: int = 500):
    """md/文本类按标题与空行拆分，约 500 字/块"""
    blocks = [b.strip() for b in re.split(r"\n\s*\n", text) if b.strip()]
    chunks, buf = [], ""
    for b in blocks:
        if buf and (len(buf) + len(b) > size or b.startswith("#")):
            chunks.append(buf)
            buf = b
        else:
            buf = (buf + "\n\n" + b) if buf else b
        while len(buf) > size * 2:  # 超长块硬切
            chunks.append(buf[:size])
            buf = buf[size:]
    if buf:
        chunks.append(buf)
    out = []
    for i, c in enumerate(chunks, 1):
        first = c.split("\n", 1)[0].strip().lstrip("#").strip()
        out.append({"seq": i, "heading": first[:40] or f"第 {i} 块", "content": c})
    return out


def _chunk_rows(headers, rows, size: int = 50):
    """表格类每 50 行一块"""
    out = []
    head_line = "列：" + " / ".join(str(h) for h in headers)
    for i in range(0, len(rows), size):
        part = rows[i:i + size]
        body = "\n".join(" | ".join(str(v) for v in r) for r in part)
        out.append({"seq": len(out) + 1, "heading": f"第 {i + 1}-{i + len(part)} 行",
                    "content": head_line + "\n" + body})
    return out or [{"seq": 1, "heading": "空表", "content": head_line}]


@router.post("/spaces/{sid}/upload")
async def upload_document(sid: int, file: UploadFile = File(...),
                          level: str = Form(default="L1"), tags: str = Form(default=""),
                          conn=Depends(db_conn), person=Depends(get_current_person)):
    """文件上传并自动解析；Excel 多工作表写 SQLite + 逐表 CSV。"""
    if person["tier"] not in UPLOAD_TIERS:
        raise HTTPException(403, "仅业务骨干/教练团/高管/开发者可上传文档")
    if not conn.execute("SELECT id FROM knowledge_spaces WHERE id=?", (sid,)).fetchone():
        raise HTTPException(404, "知识空间不存在")
    filename = file.filename or "未命名"
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(422, f"不支持的文件格式「{ext or '无扩展名'}」，"
                                 f"支持：{'/'.join(ALLOWED_EXT)}")
    raw = await file.read()
    if not raw:
        raise HTTPException(422, "文件内容为空")
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(413, "单个文件不能超过 25MB")
    level = level.upper()
    if level not in ("L1", "L2", "L3", "L4"):
        raise HTTPException(422, "密级仅支持 L1/L2/L3/L4")
    if level == "L4" and person["tier"] not in ("boss", "coach", "backbone"):
        raise HTTPException(403, "仅高管、教练团或业务骨干可上传 L4 文档")

    work_dir = UPLOAD_ROOT / f"space_{sid}"
    work_dir.mkdir(parents=True, exist_ok=True)
    title = Path(filename).stem
    did = conn.execute(
        "INSERT INTO documents(space_id,title,level,tags,uploaded_by,created_at)"
        " VALUES(?,?,?,?,?,?)",
        (sid, title, level, tags, person["name"],
         datetime.now().isoformat(timespec="seconds"))).lastrowid

    # ---- 按格式转换 + 拆分 ----
    if ext in (".txt", ".md", ".docx", ".pdf"):
        if ext == ".docx":
            text = _docx_to_text(raw)
        elif ext == ".pdf":
            text = _pdf_to_text(raw, work_dir, did)
        else:
            text = _decode(raw)
        artifact = work_dir / f"doc_{did}.md"
        artifact.write_text(text, encoding="utf-8")
        chunks = _chunk_text(text)
        fmt = "md"
    elif ext in (".csv", ".json"):
        text = _decode(raw)
        headers, rows = _csv_parse(text) if ext == ".csv" else _json_parse(text)
        db_file, cols = _rows_to_sqlite(did, headers, rows)
        artifact = work_dir / f"doc_{did}.md"
        artifact.write_text(
            f"# {title}（表格解析摘要）\n\n共 {len(rows)} 行数据，已入库 {db_file.name}。\n\n"
            + _md_table(cols, rows), encoding="utf-8")
        chunks = _chunk_rows(cols, rows)
        fmt = "sqlite"
    elif ext in (".xlsx", ".xls"):
        datasets = _excel_parse(raw, ext)
        db_file, metadata = _excel_to_sqlite_csv(did, datasets)
        total_rows = sum(item["row_count"] for item in metadata)
        artifact = work_dir / f"doc_{did}.md"
        sections, chunks = [], []
        for dataset, item in zip(datasets, metadata):
            sections.append(
                f"## {item['sheet_name']}\n\n"
                f"SQL 表：`{item['table_name']}`；{item['row_count']} 行 × "
                f"{item['column_count']} 列；CSV：`{Path(item['csv_path']).name}`。\n\n"
                + _md_table(item["columns"], dataset["rows"])
            )
            for chunk in _chunk_rows(item["columns"], dataset["rows"]):
                chunks.append({
                    "seq": len(chunks) + 1,
                    "heading": f"{item['sheet_name']} · {chunk['heading']}",
                    "content": chunk["content"],
                })
        artifact.write_text(
            f"# {title}（Excel 结构化解析摘要）\n\n"
            f"共 {len(metadata)} 个工作表、{total_rows} 行数据；已写入 {db_file.name}，"
            "并为每个工作表生成 UTF-8 CSV。\n\n" + "\n\n".join(sections),
            encoding="utf-8",
        )
        fmt = "sqlite+csv"
    else:  # .html/.htm
        cleaned = _clean_html(_decode(raw))
        artifact = work_dir / f"doc_{did}.html"
        artifact.write_text(cleaned, encoding="utf-8")
        chunks = _chunk_text(_html_to_text(cleaned))
        fmt = "html"

    for c in chunks:
        conn.execute(
            "INSERT INTO doc_chunks(document_id,seq,heading,content) VALUES(?,?,?,?)",
            (did, c["seq"], c["heading"], c["content"]))
    summary = chunks[0]["content"][:200]
    conn.execute(
        "UPDATE documents SET file_path=?, converted_format=?, chunk_count=?, summary=? WHERE id=?",
        (str(artifact), fmt, len(chunks), summary, did))
    conn.commit()
    audit(conn, person["name"], "上传解析文档", title,
          f"空间 #{sid}，格式 {ext}→{fmt}，拆分 {len(chunks)} 块")
    return dict(conn.execute("SELECT * FROM documents WHERE id=?", (did,)).fetchone())


@router.get("/documents/{did}")
def get_document(did: int, conn=Depends(db_conn), person=Depends(get_current_person)):
    row = require_document(conn, did, person)
    d = dict(row)
    d["chunks"] = [dict(r) for r in conn.execute(
        "SELECT id,seq,heading,content FROM doc_chunks WHERE document_id=? ORDER BY seq", (did,))]
    d["datasets"] = _dataset_metadata(did)
    return d


def _dataset_metadata(did: int):
    db_file = KNOWLEDGE_DB_DIR / f"doc_{did}.db"
    if not db_file.exists():
        return []
    kconn = sqlite3.connect(str(db_file))
    kconn.row_factory = sqlite3.Row
    try:
        has_index = kconn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='sheet_index'"
        ).fetchone()
        if has_index:
            return [dict(row) for row in kconn.execute(
                "SELECT sheet_name,table_name,row_count,column_count FROM sheet_index "
                "ORDER BY rowid"
            )]
        has_data = kconn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='data_rows'"
        ).fetchone()
        if not has_data:
            return []
        columns = kconn.execute("PRAGMA table_info(data_rows)").fetchall()
        count = kconn.execute("SELECT COUNT(*) FROM data_rows").fetchone()[0]
        return [{"sheet_name": "data", "table_name": "data_rows", "row_count": count,
                 "column_count": len(columns)}]
    finally:
        kconn.close()


def _require_dataset(did: int, table_name: str):
    datasets = _dataset_metadata(did)
    if table_name not in {item["table_name"] for item in datasets}:
        raise HTTPException(404, "数据表不存在")
    return KNOWLEDGE_DB_DIR / f"doc_{did}.db"


@router.get("/documents/{did}/datasets/{table_name}")
def preview_dataset(did: int, table_name: str, limit: int = 50, offset: int = 0,
                    conn=Depends(db_conn), person=Depends(get_current_person)):
    """预览解析后的 SQL 表，默认最多 50 行。"""
    require_document(conn, did, person)
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    db_file = _require_dataset(did, table_name)
    kconn = sqlite3.connect(str(db_file))
    kconn.row_factory = sqlite3.Row
    try:
        columns = [row[1] for row in kconn.execute(f'PRAGMA table_info("{table_name}")')]
        total = kconn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
        rows = [dict(row) for row in kconn.execute(
            f'SELECT * FROM "{table_name}" LIMIT ? OFFSET ?', (limit, offset)
        )]
        return {"table_name": table_name, "columns": columns, "total": total,
                "limit": limit, "offset": offset, "rows": rows}
    finally:
        kconn.close()


@router.get("/documents/{did}/database")
def download_dataset_database(did: int, conn=Depends(db_conn),
                              person=Depends(get_current_person)):
    require_document(conn, did, person)
    db_file = KNOWLEDGE_DB_DIR / f"doc_{did}.db"
    if not db_file.exists():
        raise HTTPException(404, "该文档没有 SQLite 数据库产物")
    return FileResponse(str(db_file), media_type="application/vnd.sqlite3",
                        filename=f"document_{did}.db")


@router.get("/documents/{did}/datasets/{table_name}/csv")
def download_dataset_csv(did: int, table_name: str, conn=Depends(db_conn),
                         person=Depends(get_current_person)):
    require_document(conn, did, person)
    _require_dataset(did, table_name)
    csv_file = KNOWLEDGE_DB_DIR / f"doc_{did}_csv" / f"{table_name}.csv"
    if not csv_file.exists():
        raise HTTPException(404, "该数据表没有独立 CSV 产物")
    return FileResponse(str(csv_file), media_type="text/csv; charset=utf-8",
                        filename=f"{table_name}.csv")


@router.get("/documents/{did}/file")
def get_document_file(did: int, conn=Depends(db_conn), person=Depends(get_current_person)):
    """下载/预览转换产物（.md / .html / 表格类的 .md 摘要）"""
    row = require_document(conn, did, person)
    if not row["file_path"] or not Path(row["file_path"]).exists():
        raise HTTPException(404, "该文档没有可下载的解析产物（可能为台账登记文档）")
    media = "text/html" if row["converted_format"] == "html" else "text/markdown"
    return FileResponse(row["file_path"], media_type=media,
                        filename=Path(row["file_path"]).name)
